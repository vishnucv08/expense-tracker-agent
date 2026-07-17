"""
Expense Tracker Agent
======================
Reads a CSV of bank transactions (tested with ICICI + Union Bank statement
dumps), uses Claude to categorize each transaction, keeps everything in a
pandas DataFrame in memory, and answers natural-language questions about
spending via a simple Q&A loop.

Usage:
    export ANTHROPIC_API_KEY=sk-ant-...
    python expense_tracker.py sample_transactions.csv

    # No API key handy / want to test offline?
    python expense_tracker.py sample_transactions.csv --mock

Design notes (see README section at bottom of this file / the writeup):
- Categorization: batches of transactions are sent to Claude, which returns
  JSON category labels. No hardcoded merchant->category lookup table.
- Q&A: each question is turned into a small structured "query plan" (JSON)
  by Claude, which is then executed locally against the DataFrame with
  pandas (safer + more reliable than asking the model to write/execute
  arbitrary code). A second short Claude call turns the numeric result back
  into a plain-English answer. Conversation history is kept so follow-ups
  like "what about the month before?" can be resolved.
- --mock mode swaps both Claude calls for lightweight local heuristics, so
  the whole pipeline can be exercised without network access / an API key.
"""

import argparse
import json
import re
import sys
from datetime import datetime
from dataclasses import dataclass, field

import pandas as pd

import os
import time
import hashlib
MODEL = os.environ.get("GROQ_MODEL", "llama-3.1-8b-instant")  # lighter model, higher free daily token limit than the 70b one
CATEGORIES = ["Dining", "Groceries", "Transport", "Shopping", "Bills",
              "Entertainment", "Health", "Travel", "Investments",
              "Transfers", "Loan/EMI", "Other"]
CACHE_PATH = os.environ.get("CATEGORY_CACHE", ".category_cache.json")

# --------------------------------------------------------------------------
# LLM client (lazy import so --mock mode never needs the `groq` pkg or a
# network connection). Uses Groq's free-tier API, OpenAI-compatible schema.
# --------------------------------------------------------------------------

def get_client():
    from groq import Groq
    return Groq()  # reads GROQ_API_KEY from env


def call_claude(client, system, user, max_tokens=1024, max_retries=5):
    """Name kept as call_claude for minimal diff elsewhere in the file --
    it now calls whichever LLM `get_client` set up (Groq by default).
    Retries with backoff on rate-limit errors; if Groq tells us how long
    to wait (as it does for per-minute limits), we honor that instead of
    guessing. A daily-token-limit error can't be waited out in-process --
    it's re-raised after retries so the caller can save progress and stop
    cleanly instead of looping forever."""
    try:
        from groq import RateLimitError
    except ImportError:
        RateLimitError = ()  # pragma: no cover - only relevant in --mock mode
    delay = 5
    for attempt in range(max_retries):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                max_tokens=max_tokens,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
            )
            return resp.choices[0].message.content or ""
        except RateLimitError as e:
            msg = str(e)
            # Groq's per-minute limit errors are worth retrying; a
            # tokens-per-day limit will still be exhausted after any
            # reasonable wait, so don't burn retries on that case.
            if "tokens per day" in msg.lower() or "TPD" in msg:
                raise
            if attempt == max_retries - 1:
                raise
            print(f"  Rate limited, waiting {delay}s before retry "
                  f"({attempt + 1}/{max_retries})...", file=sys.stderr)
            time.sleep(delay)
            delay = min(delay * 2, 60)


def extract_json(text):
    """Claude sometimes wraps JSON in prose or code fences; pull it out."""
    text = text.strip()
    fence = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fence:
        text = fence.group(1).strip()
    start = text.find("[") if text.lstrip().startswith("[") else text.find("{")
    end = text.rfind("]") if text.lstrip().startswith("[") else text.rfind("}")
    if start != -1 and end != -1:
        text = text[start:end + 1]
    return json.loads(text)


# --------------------------------------------------------------------------
# Step 1: Categorization
# --------------------------------------------------------------------------

def _cache_key(merchant, description):
    return hashlib.sha1(f"{merchant}|{description}".encode("utf-8")).hexdigest()


def _load_cache():
    if os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_cache(cache):
    try:
        with open(CACHE_PATH, "w") as f:
            json.dump(cache, f)
    except Exception:
        pass  # caching is a nice-to-have, never fail the run over it


def categorize_llm(df, client, batch_size=15):
    """Ask Claude to infer a category for each transaction from its
    merchant/description text. Batched to keep prompts small. Results are
    cached to disk (.category_cache.json) keyed by merchant+description,
    so re-running after a rate limit or crash doesn't re-spend tokens on
    transactions already categorized."""
    cache = _load_cache()
    categories = [None] * len(df)
    keys = [_cache_key(r.merchant, r.description) for r in df.itertuples()]

    to_fetch_idx = [i for i, k in enumerate(keys) if k not in cache]
    print(f"  {len(df) - len(to_fetch_idx)}/{len(df)} transactions already "
          f"categorized in cache; fetching {len(to_fetch_idx)} more.", file=sys.stderr)

    system = (
        "You are an expense categorization assistant. Given a list of bank "
        f"transactions, assign each one to exactly one of these categories: "
        f"{', '.join(CATEGORIES)}. Infer the category from the merchant name "
        "and description text — do not rely on any fixed lookup table, use "
        "your judgement about what kind of business/merchant this is. "
        "Real bank statements include more than everyday spending: transfers "
        "to a fixed deposit or 'TRF TO FD' -> Investments; mutual fund/SIP "
        "names (e.g. PPFAS, Bandhan) -> Investments; NACH/ECS auto-debits "
        "for a loan -> Loan/EMI; UPI payments to individual people's names "
        "(not businesses) that look like personal transfers -> Transfers. "
        "Only use Other when nothing else plausibly fits. "
        "Respond with ONLY a JSON array of category strings, one per "
        "transaction, in the same order as the input. No prose, no markdown."
    )

    try:
        for start in range(0, len(to_fetch_idx), batch_size):
            batch_positions = to_fetch_idx[start:start + batch_size]
            items = []
            for pos in batch_positions:
                r = df.iloc[pos]
                items.append({
                    "id": pos, "merchant": str(r.merchant)[:40],
                    "description": str(r.description)[:100],  # trim long UPI reference noise -- saves tokens, doesn't affect categorization
                    "amount": r.amount,
                })
            user = json.dumps(items)
            raw = call_claude(client, system, user)
            try:
                labels = extract_json(raw)
            except Exception:
                labels = ["Other"] * len(batch_positions)
            for pos, label in zip(batch_positions, labels):
                label = label if label in CATEGORIES else "Other"
                categories[pos] = label
                cache[keys[pos]] = label
            _save_cache(cache)  # persist after every batch, not just at the end
            done = start + len(batch_positions)
            print(f"  categorized {done}/{len(to_fetch_idx)} new transactions...", file=sys.stderr)
    except Exception as e:
        _save_cache(cache)
        print(f"\nStopped categorizing early ({e}). Progress saved to "
              f"{CACHE_PATH} -- re-run the same command and it'll pick up "
              f"where it left off instead of starting over.\n", file=sys.stderr)

    for i, k in enumerate(keys):
        if categories[i] is None:
            categories[i] = cache.get(k, "Other")
    return categories


def categorize_mock(df):
    """Lightweight offline stand-in for the LLM call, used with --mock.
    Keyword heuristics only — NOT a hardcoded merchant table, just enough
    to exercise the rest of the pipeline without network access."""
    rules = [
        (r"swiggy|zomato|dominos|domino's|cafe|coffee|starbucks|by2 coffee", "Dining"),
        (r"bigbasket|big bazaar|grocery|grofers", "Groceries"),
        (r"ola|uber|irctc|petrol|fuel|hp petrol|metro", "Transport"),
        (r"amazon|myntra|reliance trends|flipkart|shopping", "Shopping"),
        (r"airtel|bescom|electricity|broadband|billdesk", "Bills"),
        (r"pvr|cinema|netflix|entertainment|bookmyshow", "Entertainment"),
        (r"apollo|pharmacy|hospital|clinic|health", "Health"),
        (r"trf to fd|fixed deposit|\bppfas\b|mutual fund|\bsip\b|\bmf\b", "Investments"),
        (r"nach trxn|ecs.*loan|emi\b|loan", "Loan/EMI"),
        (r"^upi/[a-z .]+/", "Transfers"),  # person-to-person UPI with no business keyword
    ]
    out = []
    for _, row in df.iterrows():
        text = f"{row.merchant} {row.description}".lower()
        cat = "Other"
        for pattern, label in rules:
            if re.search(pattern, text):
                cat = label
                break
        out.append(cat)
    return out


def _parse_dates_robustly(raw_series):
    """Bank CSVs (especially after a round-trip through Excel) often end
    up with mixed date formats in the same column -- some rows still
    ISO 'YYYY-MM-DD', others silently reformatted to 'DD-MM-YYYY' or
    'DD/MM/YYYY' by whatever last opened the file. A single to_datetime()
    call with one implicit format drops every row that doesn't match it.
    This tries several formats per-row, in order, before giving up."""
    raw = raw_series.astype(str).str.strip()
    parsed = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")
    remaining = raw.copy()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        if remaining.empty:
            break
        attempt = pd.to_datetime(remaining, format=fmt, errors="coerce")
        newly_parsed = attempt.notna()
        parsed.loc[remaining.index[newly_parsed]] = attempt[newly_parsed]
        remaining = remaining[~newly_parsed]
    if len(remaining):
        # last resort: let pandas infer per-value, dayfirst=True since
        # this script is built around Indian bank statements
        attempt = pd.to_datetime(remaining, errors="coerce", dayfirst=True)
        newly_parsed = attempt.notna()
        parsed.loc[remaining.index[newly_parsed]] = attempt[newly_parsed]
        remaining = remaining[~newly_parsed]
    return parsed, raw.loc[remaining.index]


def load_and_categorize(csv_path, mock=False):
    df = pd.read_csv(csv_path)
    date_col_raw = df["date"]  # keep the pre-parse original for diagnostics
    df["date"], bad_raw_values = _parse_dates_robustly(df["date"])
    bad_dates = df["date"].isna().sum()
    if bad_dates:
        bad_idx = bad_raw_values.index
        originally_empty = date_col_raw.loc[bad_idx].isna()
        empty_count = int(originally_empty.sum())
        sample_rows = df.loc[bad_idx, ["merchant", "amount"]].head(5)
        sample_desc = [f"{m} (Rs. {a})" for m, a in zip(sample_rows["merchant"], sample_rows["amount"])]
        if empty_count == bad_dates:
            print(f"WARNING: {bad_dates} row(s) have a completely empty date "
                  f"cell in the CSV (not a format issue -- the cell itself is "
                  f"blank) and will be excluded from date-based views (table, "
                  f"Excel export, 'last month' style questions). These rows "
                  f"still count in overall category totals. Affected "
                  f"transactions include: {sample_desc}"
                  f"{' ...' if bad_dates > 5 else ''}", file=sys.stderr)
        else:
            samples = bad_raw_values.head(5).tolist()
            print(f"WARNING: {bad_dates} row(s) have a date that couldn't be "
                  f"parsed even after trying several common formats, and will "
                  f"be excluded from date-based views. Example raw values: "
                  f"{samples}{' ...' if bad_dates > 5 else ''}\n"
                  f"Affected transactions include: {sample_desc}"
                  f"{' ...' if bad_dates > 5 else ''}\n"
                  f"If these look like valid dates to you, share a couple of "
                  f"the raw values and it can be added.", file=sys.stderr)
        df = df[df["date"].notna()].copy()
    df["amount"] = df["amount"].astype(float)
    client = None if mock else get_client()
    df["category"] = categorize_mock(df) if mock else categorize_llm(df, client)
    return df, client


# --------------------------------------------------------------------------
# Step 2: Natural-language Q&A over the DataFrame
# --------------------------------------------------------------------------

@dataclass
class QueryPlan:
    aggregation: str = "sum"        # sum | count | average | list | max_category | breakdown | table | percentage | change
    category: str = None
    start_date: str = None
    end_date: str = None
    min_amount: float = None
    max_amount: float = None
    merchant_contains: str = None
    group_by: str = None            # "category" | "month" | None
    compare_start: str = None       # for "change": the prior period to compare against
    compare_end: str = None
    categories: list = None         # for "table": restrict to these categories (set only by deterministic override, never trusted raw from the LLM)


PLAN_SYSTEM = f"""You translate a user's question about their bank transactions
into a JSON query plan. Fields (omit any that don't apply, use null):
- aggregation: one of "sum", "count", "average", "list", "max_category", "breakdown", "table", "percentage", "change"
- category: one of {CATEGORIES} or null
- start_date / end_date: "YYYY-MM-DD" strings or null
- min_amount / max_amount: numbers or null
- merchant_contains: substring to filter merchant name, or null
- group_by: "category", "month", or null
- compare_start / compare_end: "YYYY-MM-DD" strings for the period to compare against (only for "change")

Guidance:
- "biggest expense category" / "what did I spend the most on" -> aggregation "max_category", group_by "category"
- "breakdown of spending" -> aggregation "breakdown", group_by "category"
- "month on month" / "monthly breakdown" / "table by month" / "month by month for each category" -> aggregation "table"
- "how much did I spend on X" -> aggregation "sum", category X
- "list transactions above $N" -> aggregation "list", min_amount N
- "what percentage/percent/% of my spending was on X" -> aggregation "percentage", category X (or null for all categories)
- "how much more/less did I spend in X vs Y" / "percentage change from X to Y" / "compared to last month" -> aggregation "change", start_date/end_date = the more recent period, compare_start/compare_end = the earlier period
- Today's date is {{TODAY}}. Resolve relative dates ("last month", "this month", "March") into explicit start_date/end_date.
- If the question is a follow-up (e.g. "what about the month before?"), use the
  PREVIOUS_PLAN given to you as the base and only change what the new question implies.

Respond with ONLY the JSON object, no prose, no markdown fences."""


def month_bounds(year, month):
    start = datetime(year, month, 1)
    end = datetime(year + (month == 12), (month % 12) + 1, 1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


MONTH_NAMES = ["january", "february", "march", "april", "may", "june", "july",
               "august", "september", "october", "november", "december"]
MONTH_ABBREVS = ["jan", "feb", "mar", "apr", "may", "jun", "jul",
                  "aug", "sep", "sept", "oct", "nov", "dec"]


def _find_month_mentions(q):
    """Returns [(month_number, matched_text), ...] for every month name
    OR common abbreviation (feb, sept, etc.) mentioned in the question,
    in order of appearance, using word boundaries so 'mar' doesn't match
    inside an unrelated word like 'market'."""
    hits = []
    for i, (full, abbr) in enumerate(zip(MONTH_NAMES, MONTH_ABBREVS), start=1):
        pattern = rf"\b({full}|{abbr})\b"
        m = re.search(pattern, q)
        if m:
            hits.append((m.start(), i, m.group(1)))
    hits.sort(key=lambda h: h[0])  # order of appearance in the question
    return [(i, text) for _, i, text in hits]


def _find_month_year_mentions(q):
    """Returns [(month_number, year), ...] for explicit 'month year'
    pairs like 'july 2025' or 'feb 2026', in order of appearance. Used
    for full date ranges spanning specific years, e.g. 'from July 2025
    to July 2026' -- plain month-name detection can't tell those two
    "julys" apart."""
    all_months = "|".join(MONTH_NAMES + MONTH_ABBREVS)
    pattern = rf"\b({all_months})\.?\s+(\d{{4}})\b"
    hits = []
    for m in re.finditer(pattern, q):
        token = m.group(1)
        year = int(m.group(2))
        month_num = (MONTH_NAMES.index(token) + 1 if token in MONTH_NAMES
                     else MONTH_ABBREVS.index(token) + 1)
        hits.append((m.start(), month_num, year))
    hits.sort(key=lambda h: h[0])
    return [(mo, y) for _, mo, y in hits]


def resolve_date_cues(question, today, previous_plan):
    """Deterministically figure out date-related plan fields from the
    question text -- 'last month', 'the month before', bare month names
    (full or abbreviated, e.g. 'feb'), explicit 'month year' ranges (e.g.
    'July 2025 to July 2026'), and 'X vs Y' / 'X compared to Y' two-month
    comparisons. Returns only the keys it's confident about (empty dict
    if no date cue found at all), so callers can force these onto a plan
    regardless of what an LLM guessed. This exists because small/fast
    LLMs are unreliable at date arithmetic -- they've been observed
    inventing wrong years and zero-ing out ranges that actually have
    data."""
    q = question.lower()
    today_dt = datetime.strptime(today, "%Y-%m-%d")
    out = {}

    year_month_hits = _find_month_year_mentions(q)
    is_change_phrasing_early = any(w in q for w in
                                    ("vs", "versus", "compared to", "compare",
                                     "increase", "decrease", "change from"))

    if "last month" in q:
        m = today_dt.month - 1 or 12
        y = today_dt.year if today_dt.month > 1 else today_dt.year - 1
        out["start_date"], out["end_date"] = month_bounds(y, m)
    elif "month before" in q or "before that" in q or "previous month" in q:
        base_start = (previous_plan or {}).get("start_date")
        if not base_start:
            m = today_dt.month - 1 or 12
            y = today_dt.year if today_dt.month > 1 else today_dt.year - 1
            base_start, _ = month_bounds(y, m)
        d = datetime.strptime(base_start, "%Y-%m-%d")
        m = d.month - 1 or 12
        y = d.year if d.month > 1 else d.year - 1
        out["start_date"], out["end_date"] = month_bounds(y, m)
    elif len(year_month_hits) >= 2 and is_change_phrasing_early:
        # explicit years, comparison phrasing -> "change" between two
        # specific (month, year) periods, e.g. "March 2026 vs Feb 2026"
        (m1, y1), (m2, y2) = year_month_hits[0], year_month_hits[1]
        chrono = sorted([(y1, m1), (y2, m2)])
        out["compare_start"], out["compare_end"] = month_bounds(chrono[0][0], chrono[0][1])
        out["start_date"], out["end_date"] = month_bounds(chrono[1][0], chrono[1][1])
    elif len(year_month_hits) >= 2:
        # explicit years, no comparison phrasing -> a full date RANGE,
        # e.g. "from July 2025 to July 2026" spans every month in between
        (m1, y1), (m2, y2) = year_month_hits[0], year_month_hits[-1]
        chrono = sorted([(y1, m1), (y2, m2)])
        out["start_date"], _ = month_bounds(chrono[0][0], chrono[0][1])
        _, out["end_date"] = month_bounds(chrono[1][0], chrono[1][1])
    elif len(year_month_hits) == 1 and any(
            i != year_month_hits[0][0] for i, _ in _find_month_mentions(q)):
        # exactly one month has an explicit year, but another bare month
        # is also mentioned -- e.g. "feb to june 2026" (feb has no year
        # of its own). Assume the bare month shares the explicit year;
        # this is the common case for a same-year range.
        explicit_m, explicit_y = year_month_hits[0]
        other = next(i for i, _ in _find_month_mentions(q) if i != explicit_m)
        pairs = sorted([(explicit_y, explicit_m), (explicit_y, other)])
        if is_change_phrasing_early:
            out["compare_start"], out["compare_end"] = month_bounds(*pairs[0])
            out["start_date"], out["end_date"] = month_bounds(*pairs[1])
        else:
            out["start_date"], _ = month_bounds(*pairs[0])
            _, out["end_date"] = month_bounds(*pairs[1])
    elif len(year_month_hits) == 1:
        # a single explicit "month year" -> that exact month, no guessing needed
        m, y = year_month_hits[0]
        out["start_date"], out["end_date"] = month_bounds(y, m)
    else:
        month_hits = _find_month_mentions(q)
        is_change_phrasing = any(w in q for w in
                                  ("vs", "versus", "compared to", "compare", "more than last",
                                   "less than last", "increase", "decrease", "change from"))
        if is_change_phrasing and len(month_hits) >= 2:
            (i1, _), (i2, _) = month_hits[0], month_hits[1]
            first, second = sorted([i1, i2])
            # bare month name -> most recent occurrence not in the future
            y1 = today_dt.year if first <= today_dt.month else today_dt.year - 1
            y2 = today_dt.year if second <= today_dt.month else today_dt.year - 1
            out["compare_start"], out["compare_end"] = month_bounds(y1, first)
            out["start_date"], out["end_date"] = month_bounds(y2, second)
        elif is_change_phrasing and ("last month" in q or "compared to last" in q or "vs last month" in q):
            m_cur = today_dt.month
            y_cur = today_dt.year
            out["start_date"], out["end_date"] = month_bounds(y_cur, m_cur)
            m_prev = m_cur - 1 or 12
            y_prev = y_cur if m_cur > 1 else y_cur - 1
            out["compare_start"], out["compare_end"] = month_bounds(y_prev, m_prev)
        elif month_hits:
            i, _ = month_hits[0]
            # bare month name with no year -> the most recent occurrence
            # of that month that isn't in the future (not always "this
            # year" -- e.g. asking about "December" in July should mean
            # last December, not five months from now)
            y = today_dt.year if i <= today_dt.month else today_dt.year - 1
            out["start_date"], out["end_date"] = month_bounds(y, i)
    return out


def get_plan_llm(client, question, today, previous_plan):
    system = PLAN_SYSTEM.replace("{TODAY}", today)
    user = f"PREVIOUS_PLAN: {json.dumps(previous_plan)}\nQUESTION: {question}"
    raw = call_claude(client, system, user, max_tokens=400)
    data = extract_json(raw)
    plan = QueryPlan(**{k: data.get(k) for k in QueryPlan.__dataclass_fields__})
    # Defensive: the LLM was observed returning a LIST for `category` when
    # a question mentions more than one (e.g. "dining and shopping"),
    # even though the plan schema expects a single string. Anything
    # downstream that calls plan.category.lower() would crash on a list.
    # Coerce to a single valid category (first match) or None -- true
    # multi-category support for "table" requests is handled separately
    # via plan.categories, set deterministically below.
    if isinstance(plan.category, (list, tuple)):
        valid = [c for c in plan.category if isinstance(c, str) and c in CATEGORIES]
        plan.category = valid[0] if valid else None
    elif plan.category is not None and not isinstance(plan.category, str):
        plan.category = None
    # Hard override: don't trust the LLM's date arithmetic (observed
    # inventing wrong years and zero-ing out ranges that have real data).
    # Deterministic parsing wins whenever the question has a date cue.
    for k, v in resolve_date_cues(question, today, previous_plan).items():
        setattr(plan, k, v)
    # Hard override: on a follow-up ("and what about the month before?"),
    # the LLM correctly carries the date range forward but was observed
    # silently dropping the category filter from the previous turn,
    # e.g. answering with total spend instead of "Dining" spend. If this
    # looks like a follow-up, the new question doesn't name a different
    # category, and the LLM's plan has no category, carry the previous
    # one forward.
    q = question.lower()
    if (previous_plan and any(cue in q for cue in FOLLOWUP_CUES)
            and plan.category is None
            and not any(cat.lower() in q for cat in CATEGORIES)
            and previous_plan.get("category")):
        plan.category = previous_plan["category"]
    return plan


def _normalize_text(s):
    """Strip everything but letters/digits, lowercased -- for comparing a
    merchant name against question text regardless of spacing/punctuation
    differences (e.g. 'By2Coffee' in a question vs 'By2 Coffee' in the
    data)."""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def resolve_merchant_mention(question, df):
    """If the question text contains something that matches an actual
    merchant name in the data, return that exact merchant string. This
    is far more reliable than trusting the LLM to extract a fuzzy
    merchant name into merchant_contains on its own -- observed failure
    modes were matching nothing (exact-text/spacing mismatch) and
    matching far too much (silently dropping the filter and effectively
    summing almost the whole dataset). Comparison is done on normalized
    text (letters/digits only) so spacing and punctuation don't matter."""
    q_norm = _normalize_text(question)
    if not q_norm:
        return None
    merchants = df["merchant"].dropna().unique()
    matches = []
    for m in merchants:
        m_norm = _normalize_text(m)
        if m_norm and len(m_norm) >= 3 and m_norm in q_norm:
            matches.append(m)
    if not matches:
        return None
    # prefer the longest match -- avoids a short merchant name
    # incorrectly matching as a substring of a different, longer one
    return max(matches, key=lambda m: len(_normalize_text(m)))


FOLLOWUP_CUES = ("what about", "and the", "before that", "month before",
                  "previous month", "instead", "same for", "how about")


def get_plan_mock(question, today, previous_plan):
    """Offline heuristic query planner for --mock mode."""
    q = question.lower()
    is_followup = previous_plan and any(cue in q for cue in FOLLOWUP_CUES)
    plan = dict(previous_plan) if is_followup else {}
    plan.setdefault("aggregation", "sum")

    for cat in CATEGORIES:
        if cat.lower() in q:
            plan["category"] = cat
    if not is_followup and not any(cat.lower() in q for cat in CATEGORIES):
        plan["category"] = None

    today_dt = datetime.strptime(today, "%Y-%m-%d")
    plan.update(resolve_date_cues(question, today, previous_plan))

    m = re.search(r"(?:above|over|more than|greater than)\s*\$?(\d+)", q)
    if m:
        plan["min_amount"] = float(m.group(1))
        plan["aggregation"] = "list"
    m = re.search(r"(?:below|under|less than)\s*\$?(\d+)", q)
    if m:
        plan["max_amount"] = float(m.group(1))
        plan["aggregation"] = "list"

    # "X vs Y" / "X compared to Y" -> a "change" comparison
    is_change_phrasing = any(w in q for w in
                              ("vs", "versus", "compared to", "compare", "more than last",
                               "less than last", "increase", "decrease", "change from"))

    if "percent" in q or "%" in q:
        plan["aggregation"] = "percentage"
    elif is_change_phrasing:
        plan["aggregation"] = "change"
        # start_date/end_date/compare_start/compare_end already set
        # deterministically by resolve_date_cues() above.
    elif "table" in q or "month on month" in q or "month-on-month" in q or "month by month" in q:
        plan["aggregation"] = "table"
    elif "biggest" in q or "most" in q or "highest" in q:
        plan["aggregation"] = "max_category"
        plan["group_by"] = "category"
    elif "breakdown" in q or "by category" in q:
        plan["aggregation"] = "breakdown"
        plan["group_by"] = "category"
    elif "list" in q or "show" in q:
        plan["aggregation"] = "list"
    elif "average" in q or "avg" in q:
        plan["aggregation"] = "average"
    elif "how many" in q or "count" in q:
        plan["aggregation"] = "count"

    return QueryPlan(**{k: plan.get(k) for k in QueryPlan.__dataclass_fields__})


def _safe_category_str(category):
    """Defensive: plan.category should always be a string or None by the
    time it reaches here (get_plan_llm coerces it), but this is a last
    line of defense against a crash if that ever slips -- e.g. a raw
    list would crash on .lower() otherwise."""
    if category is None:
        return None
    if isinstance(category, (list, tuple)):
        return str(category[0]) if category else None
    return str(category)


def _apply_non_date_filters(df, plan):
    """category/amount/merchant filters only -- used both for the main
    query and for building comparison periods without re-filtering by
    category twice."""
    out = df.copy()
    cat = _safe_category_str(plan.category)
    if cat:
        out = out[out["category"].str.lower() == cat.lower()]
    if plan.min_amount is not None:
        out = out[out["amount"] > plan.min_amount]
    if plan.max_amount is not None:
        out = out[out["amount"] < plan.max_amount]
    if plan.merchant_contains:
        out = out[out["merchant"].str.contains(plan.merchant_contains, case=False, na=False)]
    return out


def apply_plan(df, plan: QueryPlan):
    out = df.copy()
    cat = _safe_category_str(plan.category)
    if cat:
        out = out[out["category"].str.lower() == cat.lower()]
    if plan.start_date:
        out = out[out["date"] >= plan.start_date]
    if plan.end_date:
        out = out[out["date"] < plan.end_date]
    if plan.min_amount is not None:
        out = out[out["amount"] > plan.min_amount]
    if plan.max_amount is not None:
        out = out[out["amount"] < plan.max_amount]
    if plan.merchant_contains:
        out = out[out["merchant"].str.contains(plan.merchant_contains, case=False, na=False)]

    if plan.aggregation == "sum":
        return {"total": round(out["amount"].sum(), 2), "n": len(out)}
    if plan.aggregation == "count":
        return {"count": len(out)}
    if plan.aggregation == "average":
        return {"average": round(out["amount"].mean(), 2) if len(out) else 0, "n": len(out)}
    if plan.aggregation == "list":
        rows = out.sort_values("amount", ascending=False)
        return {"transactions": rows[["date", "merchant", "amount", "category"]]
                .assign(date=lambda d: d["date"].dt.strftime("%Y-%m-%d"))
                .to_dict("records")}
    if plan.aggregation in ("max_category", "breakdown"):
        grouped = out.groupby("category")["amount"].sum().sort_values(ascending=False)
        result = {"by_category": {k: round(v, 2) for k, v in grouped.items()}}
        if plan.aggregation == "max_category" and len(grouped):
            result["top_category"] = grouped.index[0]
            result["top_amount"] = round(grouped.iloc[0], 2)
        return result
    if plan.aggregation == "percentage":
        # Denominator = everything in the date/amount/merchant range,
        # ACROSS all categories (not just plan.category) -- that's what
        # "percentage of spending" means. Numerator = the category-filtered
        # subset (the same `out` computed above, which already has the
        # category filter applied if plan.category was set).
        denom_df = df.copy()
        if plan.start_date:
            denom_df = denom_df[denom_df["date"] >= plan.start_date]
        if plan.end_date:
            denom_df = denom_df[denom_df["date"] < plan.end_date]
        if plan.min_amount is not None:
            denom_df = denom_df[denom_df["amount"] > plan.min_amount]
        if plan.max_amount is not None:
            denom_df = denom_df[denom_df["amount"] < plan.max_amount]
        total = denom_df["amount"].sum()
        if plan.category:
            cat_total = out["amount"].sum()
            pct = round(cat_total / total * 100, 2) if total else 0
            return {"percentage": pct, "category_total": round(cat_total, 2), "grand_total": round(total, 2)}
        grouped = denom_df.groupby("category")["amount"].sum().sort_values(ascending=False)
        pct_by_cat = {k: round(v / total * 100, 2) if total else 0 for k, v in grouped.items()}
        return {"percentage_by_category": pct_by_cat, "grand_total": round(total, 2)}
    if plan.aggregation == "change":
        current_filtered = _apply_non_date_filters(df, plan)
        current = current_filtered.copy()
        if plan.start_date:
            current = current[current["date"] >= plan.start_date]
        if plan.end_date:
            current = current[current["date"] < plan.end_date]
        prior = current_filtered.copy()
        if plan.compare_start:
            prior = prior[prior["date"] >= plan.compare_start]
        if plan.compare_end:
            prior = prior[prior["date"] < plan.compare_end]
        cur_total = round(current["amount"].sum(), 2)
        prior_total = round(prior["amount"].sum(), 2)
        diff = round(cur_total - prior_total, 2)
        pct_change = round(diff / prior_total * 100, 2) if prior_total else None
        return {"current_total": cur_total, "prior_total": prior_total,
                "diff": diff, "pct_change": pct_change,
                "current_range": [plan.start_date, plan.end_date],
                "prior_range": [plan.compare_start, plan.compare_end]}
    if plan.aggregation == "table":
        # Table ignores the row-level date/category filters computed into
        # `out` above (a month-on-month table needs its own month-by-month
        # breakdown, not a single pre-filtered slice) -- but DOES respect
        # an explicit category subset (plan.categories) or date range
        # (plan.start_date/end_date) if the question named one, e.g.
        # "table of dining and shopping from Feb to June 2026".
        months = (month_range_between(plan.start_date, plan.end_date)
                  if plan.start_date and plan.end_date else None)
        table = category_month_table(df, months=months, categories=plan.categories)
        label = (f"{months[0][0]} - {months[-1][0]}" if months
                 else month_range_label(df))
        return {"table_text": format_table_grid(table),
                "range_label": label}
    return {"total": round(out["amount"].sum(), 2), "n": len(out)}


ANSWER_SYSTEM = """You are a helpful personal finance assistant. Given the
user's question, the structured query plan that was executed, and the
numeric result, write a short, natural, plain-English answer (1-3
sentences). Use currency figures as given (assume INR, prefix with Rs.).
Don't mention "query plan" or JSON - just answer like a helpful assistant."""


def format_answer_llm(client, question, plan, result):
    user = (f"QUESTION: {question}\nPLAN: {json.dumps(plan.__dict__)}\n"
            f"RESULT: {json.dumps(result, default=str)}")
    return call_claude(client, ANSWER_SYSTEM, user, max_tokens=300).strip()


def format_answer_mock(question, plan, result):
    if "table_text" in result:
        return f"Month-on-month spending by category ({result.get('range_label','')}):\n" + result["table_text"]
    if "total" in result:
        scope = f" on {plan.category}" if plan.category else ""
        scope += f" at {plan.merchant_contains}" if plan.merchant_contains else ""
        when = f" between {plan.start_date} and {plan.end_date}" if plan.start_date else ""
        return f"You spent Rs. {result['total']:,.2f}{scope}{when} across {result['n']} transaction(s)."
    if "count" in result:
        return f"There were {result['count']} matching transaction(s)."
    if "average" in result:
        return f"The average was Rs. {result['average']:,.2f} across {result['n']} transaction(s)."
    if "transactions" in result:
        rows = result["transactions"]
        if not rows:
            return "No transactions matched that."
        shown = rows[:30]
        lines = [f"  - {r['date']}  {r['merchant']:<20} Rs. {r['amount']:.2f}  [{r['category']}]" for r in shown]
        header = f"Matching transactions ({len(rows)} total"
        header += f", showing top {len(shown)} by amount" if len(rows) > 30 else ""
        header += "):"
        return header + "\n" + "\n".join(lines)
    if "top_category" in result:
        return (f"Your biggest spending category was {result['top_category']} "
                f"at Rs. {result['top_amount']:,.2f}.")
    if "by_category" in result:
        lines = [f"  - {k}: Rs. {v:,.2f}" for k, v in result["by_category"].items()]
        return "Spending breakdown:\n" + "\n".join(lines)
    if "percentage" in result:
        scope = f" on {plan.category}" if plan.category else ""
        return (f"You spent Rs. {result['category_total']:,.2f}{scope}, which is "
                f"{result['percentage']:.1f}% of your total spend of Rs. {result['grand_total']:,.2f}"
                f" in that range.")
    if "percentage_by_category" in result:
        lines = [f"  - {k}: {v:.1f}%" for k, v in result["percentage_by_category"].items()]
        return (f"Percentage of spending by category (of Rs. {result['grand_total']:,.2f} total):\n"
                + "\n".join(lines))
    if "pct_change" in result:
        direction = "more" if result["diff"] >= 0 else "less"
        pct_text = f"{abs(result['pct_change']):.1f}%" if result["pct_change"] is not None else "N/A (no prior spend to compare)"
        return (f"You spent Rs. {result['current_total']:,.2f} in the more recent period vs "
                f"Rs. {result['prior_total']:,.2f} before -- that's Rs. {abs(result['diff']):,.2f} "
                f"{direction} ({pct_text}).")
    return "I couldn't compute an answer for that."


# --------------------------------------------------------------------------
# Stretch: budget alerts + simple chart
# --------------------------------------------------------------------------

def budget_alerts(df, budgets: dict):
    """budgets: {"Dining": 3000, "Shopping": 5000, ...} -> flags categories
    (over all loaded data) that exceed the given limit."""
    alerts = []
    totals = df.groupby("category")["amount"].sum()
    for cat, limit in budgets.items():
        spent = totals.get(cat, 0)
        if spent > limit:
            alerts.append(f"⚠️  {cat}: Rs. {spent:,.2f} spent, over budget of Rs. {limit:,.2f} "
                           f"(by Rs. {spent - limit:,.2f})")
    return alerts


def make_chart(df, out_path="spending_by_category.png"):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    totals = df.groupby("category")["amount"].sum().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(8, 5))
    totals.plot(kind="bar", ax=ax, color="#4C72B0")
    ax.set_ylabel("Amount (Rs.)")
    ax.set_title("Spending by Category")
    plt.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def default_feb_to_jun_months(df):
    """(label, start, end) tuples for Feb-Jun of the data's own year.
    Kept for backwards compatibility / explicit use; auto_month_range()
    below is the actual default now since real statements can span any
    range of months."""
    year = df["date"].dt.year.min()
    month_names = ["Feb", "Mar", "Apr", "May", "Jun"]
    months = []
    for i, name in enumerate(month_names, start=2):
        start, end = month_bounds(int(year), i)
        months.append((name, start, end))
    return months


def auto_month_range(df):
    """(label, start, end) tuples for every calendar month actually
    present in the data, in chronological order. This is the real
    default: statements can span any date range, not just Feb-Jun."""
    valid_dates = df["date"].dropna()
    periods = sorted(valid_dates.dt.to_period("M").unique())
    months = []
    for p in periods:
        label = p.strftime("%b %Y")
        start = p.start_time.strftime("%Y-%m-%d")
        end = (p + 1).start_time.strftime("%Y-%m-%d")
        months.append((label, start, end))
    return months


def month_range_label(df):
    months = auto_month_range(df)
    if not months:
        return ""
    return f"{months[0][0]} - {months[-1][0]}"


def format_table_grid(table):
    """Render a category-by-month DataFrame as a proper bordered grid
    instead of pandas' plain-text dump (no alignment guides otherwise)."""
    try:
        from tabulate import tabulate
        display = table.reset_index()
        # Pass raw numbers with floatfmt, not pre-formatted strings --
        # tabulate re-parses numeric-looking strings itself and mangles
        # large values into scientific notation if you hand it text.
        return tabulate(display, headers="keys", tablefmt="grid",
                         showindex=False, floatfmt=",.2f")
    except ImportError:
        return table.to_string(float_format=lambda x: f"{x:,.2f}")


def month_range_between(start_date, end_date):
    """(label, start, end) tuples for every calendar month from start_date
    up to (but not including) end_date -- used when a table request names
    an explicit range, e.g. 'Feb to June 2026', instead of defaulting to
    every month present in the data. end_date is exclusive, consistent
    with the rest of the codebase (it's the start of the month AFTER the
    last one to include)."""
    start = pd.Period(start_date, freq="M")
    end_inclusive = pd.Period(end_date, freq="M") - 1
    months = []
    p = start
    while p <= end_inclusive:
        label = p.strftime("%b %Y")
        m_start = p.start_time.strftime("%Y-%m-%d")
        m_end = (p + 1).start_time.strftime("%Y-%m-%d")
        months.append((label, m_start, m_end))
        p += 1
    return months


def category_month_table(df, months=None, categories=None):
    """Return a pandas DataFrame: categories as rows, month labels as
    columns (plus a Total column), values = spend in that category/month.
    `categories`, if given, restricts the rows to just that subset (e.g.
    a question asking about "dining and shopping" specifically)."""
    if months is None:
        months = auto_month_range(df)
    if categories:
        df = df[df["category"].isin(categories)]
        categories_present = [c for c in categories if c in set(df["category"])]
    else:
        categories_present = sorted(df["category"].unique())
    table = pd.DataFrame(index=categories_present, columns=[m[0] for m in months], dtype=float)
    for label, start, end in months:
        mask = (df["date"] >= start) & (df["date"] < end)
        totals = df[mask].groupby("category")["amount"].sum()
        table[label] = table.index.map(lambda c: round(totals.get(c, 0.0), 2))
    table["Total"] = table.sum(axis=1).round(2)
    table.loc["Total"] = table.sum(axis=0).round(2)
    table.index.name = "Category"
    return table


def export_excel(df, out_path="spending_by_month.xlsx", months=None):
    """Write an .xlsx with two sheets:
      - Transactions: the raw categorized data (one row per transaction)
      - Summary: categories as rows, months as columns, each cell a SUMIFS
        formula over the Transactions sheet (not a hardcoded number), so it
        recalculates if you edit/add rows in Excel later.
    `months` is a list of (label, "YYYY-MM-DD" start, "YYYY-MM-DD" end)
    tuples; defaults to every month present in the data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    if months is None:
        months = auto_month_range(df)

    categories_present = sorted(df["category"].unique())

    wb = Workbook()

    # --- Transactions sheet (raw data SUMIFS will read from) ---
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    headers = ["Date", "Bank", "Merchant", "Description", "Amount", "Category"]
    ws_tx.append(headers)
    for cell in ws_tx[1]:
        cell.font = Font(name="Arial", bold=True)
    for _, row in df.sort_values("date").iterrows():
        ws_tx.append([
            row["date"].to_pydatetime(),
            row.get("bank", ""),
            row["merchant"],
            row["description"],
            float(row["amount"]),
            row["category"],
        ])
    for col_idx, width in zip(range(1, 7), [12, 12, 18, 32, 12, 14]):
        ws_tx.column_dimensions[get_column_letter(col_idx)].width = width
    for r in range(2, ws_tx.max_row + 1):
        ws_tx.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws_tx.cell(row=r, column=5).number_format = "#,##0.00"
    n_tx = ws_tx.max_row  # includes header row

    # --- Summary sheet: categories as rows, months as columns ---
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Category"
    ws["A1"].font = Font(name="Arial", bold=True)
    for j, (label, _, _) in enumerate(months, start=2):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name="Arial", bold=True)
        c.alignment = Alignment(horizontal="center")
    total_col = len(months) + 2
    ws.cell(row=1, column=total_col, value="Total").font = Font(name="Arial", bold=True)

    amount_range = f"Transactions!$E$2:$E${n_tx}"
    category_range = f"Transactions!$F$2:$F${n_tx}"
    date_range = f"Transactions!$A$2:$A${n_tx}"

    for i, cat in enumerate(categories_present, start=2):
        ws.cell(row=i, column=1, value=cat).font = Font(name="Arial")
        for j, (label, start, end) in enumerate(months, start=2):
            col = get_column_letter(j)
            formula = (f'=SUMIFS({amount_range},{category_range},$A{i},'
                       f'{date_range},">={start}",{date_range},"<{end}")')
            cell = ws.cell(row=i, column=j, value=formula)
            cell.number_format = "#,##0.00"
        total_formula = f"=SUM({get_column_letter(2)}{i}:{get_column_letter(total_col - 1)}{i})"
        tcell = ws.cell(row=i, column=total_col, value=total_formula)
        tcell.number_format = "#,##0.00"
        tcell.font = Font(name="Arial", bold=True)

    last_row = len(categories_present) + 2
    ws.cell(row=last_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    for j in range(2, total_col + 1):
        col = get_column_letter(j)
        f = f"=SUM({col}2:{col}{last_row - 1})"
        c = ws.cell(row=last_row, column=j, value=f)
        c.number_format = "#,##0.00"
        c.font = Font(name="Arial", bold=True)

    ws.column_dimensions["A"].width = 16
    for j in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12

    wb.save(out_path)
    return out_path




def main():
    parser = argparse.ArgumentParser(description="Expense Tracker Agent")
    parser.add_argument("csv_path")
    parser.add_argument("--mock", action="store_true",
                         help="Run fully offline with heuristic stand-ins instead of Claude API calls")
    parser.add_argument("--chart", action="store_true", help="Save a spending-by-category bar chart on exit")
    parser.add_argument("--table", action="store_true",
                         help="Print a category-by-month table (all months present in the data) to the terminal")
    parser.add_argument("--excel", nargs="?", const="spending_by_month.xlsx", default=None,
                         help="Save a categories-x-months Excel summary (default: spending_by_month.xlsx)")
    parser.add_argument("--budget", action="append", default=[],
                         help='Budget limit, e.g. --budget Dining=3000 (repeatable)')
    parser.add_argument("--questions", nargs="*", default=None,
                         help="Run non-interactively with this list of questions instead of an input loop")
    args = parser.parse_args()

    print(f"Loading {args.csv_path} and categorizing transactions "
          f"({'mock/offline' if args.mock else 'via LLM API'})...")
    df, client = load_and_categorize(args.csv_path, mock=args.mock)
    print(f"Loaded {len(df)} transactions across "
          f"{df['date'].dt.to_period('M').nunique()} months.\n")
    print(df.groupby("category")["amount"].sum().sort_values(ascending=False).to_string())
    print()

    budgets = {}
    for b in args.budget:
        cat, val = b.split("=")
        budgets[cat] = float(val)
    if budgets:
        alerts = budget_alerts(df, budgets)
        for a in alerts:
            print(a)
        if not alerts:
            print("No budget alerts — all categories within limits.")
        print()

    if args.chart:
        path = make_chart(df)
        print(f"Saved chart to {path}\n")

    if args.table:
        table = category_month_table(df)
        print(f"Month-on-month spending by category ({month_range_label(df)}):")
        print(format_table_grid(table))
        print()

    if args.excel:
        path = export_excel(df, args.excel)
        print(f"Saved Excel summary to {path}\n")

    today = datetime.now().strftime("%Y-%m-%d")
    previous_plan = {}

    TABLE_CUES = ("month on month", "month-on-month", "month by month",
                  "monthly table", "monthly breakdown", "table for each category",
                  "table by category", "give me a table", "as a table", "table")

    def handle(question):
        nonlocal previous_plan
        if args.mock:
            plan = get_plan_mock(question, today, previous_plan)
        else:
            plan = get_plan_llm(client, question, today, previous_plan)
        # Hard override: don't rely on the planner alone to route table
        # requests correctly (it can pick "breakdown" instead). If the
        # question clearly asks for a month-by-month table, force it.
        if any(cue in question.lower() for cue in TABLE_CUES):
            plan.aggregation = "table"
            # If the question names specific categories (e.g. "table of
            # dining and shopping"), restrict the table to just those
            # instead of every category. This is also what fixed a real
            # crash: the LLM was returning a LIST for plan.category on
            # multi-category questions, which broke the single-category
            # filter elsewhere -- categories are only ever read here, for
            # the table, never assigned back onto plan.category.
            mentioned = [c for c in CATEGORIES if c.lower() in question.lower()]
            if mentioned:
                plan.categories = mentioned
        # Hard override: match merchant names against the real data
        # instead of trusting the LLM's guessed substring -- observed
        # matching nothing (spacing mismatch) or matching almost
        # everything (filter silently dropped). On a follow-up with no
        # merchant mentioned in the new question, carry the previous
        # turn's merchant filter forward instead of losing it.
        merchant_hit = resolve_merchant_mention(question, df)
        if merchant_hit:
            plan.merchant_contains = merchant_hit
        elif (previous_plan and any(cue in question.lower() for cue in FOLLOWUP_CUES)
              and previous_plan.get("merchant_contains")):
            plan.merchant_contains = previous_plan["merchant_contains"]
        result = apply_plan(df, plan)
        if "table_text" in result:
            answer = f"Month-on-month spending by category ({result.get('range_label','')}):\n" + result["table_text"]
        else:
            # Always format deterministically from the computed result dict,
            # even in live (non-mock) mode. Letting the LLM restate exact
            # numbers in prose (format_answer_llm) was observed hallucinating
            # figures -- e.g. reporting Rs. 20,525 for a category/month
            # combination that actually totalled Rs. 5,338.75. The numbers
            # are already computed correctly by pandas; there's no reason to
            # let a second LLM call risk corrupting them on the way out.
            answer = format_answer_mock(question, plan, result)
        previous_plan = plan.__dict__
        return answer

    if args.questions:
        for q in args.questions:
            print(f"> {q}")
            print(handle(q))
            print()
        return

    print("Ask questions about your spending (type 'quit' to exit).")
    while True:
        try:
            q = input("> ").strip()
        except EOFError:
            break
        if not q or q.lower() in ("quit", "exit"):
            break
        try:
            print(handle(q))
        except Exception as e:
            print(f"Sorry, I hit an error answering that: {e}")
        print()


if __name__ == "__main__":
    main()
